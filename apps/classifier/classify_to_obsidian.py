#!/usr/bin/env python3
import argparse
import base64
import hashlib
import json
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import time
import zipfile
from collections import Counter
from datetime import datetime
from json import JSONDecoder
from pathlib import Path, PurePosixPath
from typing import Any, Dict, Iterable, List, Optional
from urllib.parse import quote
from xml.etree import ElementTree as ET
from zoneinfo import ZoneInfo

import requests
from packages.classification.ocr_pipeline import (
    extract_image_text_with_metadata,
    extract_pdf_text_with_metadata,
)
from packages.classification.retrieval_metadata import build_retrieval_metadata
from packages.runtime import load_classifier_runtime_settings
from packages.vault.naming import (
    build_attachment_filename,
    build_extracted_markdown_filename,
    build_note_filename,
)

from .codex_arbiter import run_codex_final_arbiter
from .category_manager import (
    find_reviewed_label_override,
    format_examples_for_prompt,
    image_safe_categories,
    is_text_message_source_path,
    load_categories,
    load_relevant_examples,
    normalize_image_classification_result,
    select_candidate_categories,
)
from .hybrid_runtime import (
    LIGHTGBM_MODEL_PATH,
    choose_live_decision,
    enqueue_shadow_job,
    ensure_lightgbm_model,
    load_hybrid_gating_config,
    load_heuristic_rules,
    maybe_retrain_from_shadow_data,
    predict_lightgbm_result,
    run_autonomous_shadow_cycle,
    write_readiness_report,
)
from .ollama_runtime import wait_for_ollama

ALASKA_TZ = ZoneInfo("America/Anchorage")
RUNTIME_SETTINGS = load_classifier_runtime_settings()

SUPPORTED_EXTENSIONS = {
    ".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt",
    ".txt", ".md", ".markdown", ".csv", ".html", ".htm",
    ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"
}

IMAGE_EXTENSIONS = {
    ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"
}

SPREADSHEET_EXTENSIONS = {
    ".xlsx"
}

DOCX_EXTENSIONS = {
    ".docx"
}

PDF_EXTENSIONS = {
    ".pdf"
}

PLAIN_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".csv"
}

IMAGE_OCR_MIN_CHARS = int(os.environ.get("IMAGE_OCR_MIN_CHARS", "48") or "48")


def extraction_quality_for_text(text: str) -> str:
    chars = sum(1 for char in str(text or "") if char.isalnum())
    if chars >= 120:
        return "high"
    if chars >= 20:
        return "medium"
    if chars >= 8:
        return "low"
    return "empty"

def now_ak() -> str:
    return datetime.now(ALASKA_TZ).isoformat(timespec="seconds")

def elapsed_ms(started_at: float) -> float:
    return round((time.perf_counter() - started_at) * 1000, 3)

def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def slugify(value: str, max_len: int = 100) -> str:
    value = re.sub(r"[^A-Za-z0-9._ -]+", "", value)
    value = re.sub(r"\s+", " ", value).strip()
    value = value.replace("/", "-")
    return value[:max_len] or "document"

def obsidian_tag(value: str) -> str:
    value = str(value).lower().strip()
    value = re.sub(r"[^a-z0-9_-]+", "-", value)
    return value.strip("-") or "unknown"


def _coerce_secondary_labels(raw_value: Any) -> List[str]:
    if isinstance(raw_value, (list, tuple, set)):
        return [str(item) for item in raw_value if str(item).strip()]
    if isinstance(raw_value, str) and raw_value.strip():
        return [raw_value.strip()]
    return []


def _coerce_float(raw_value: Any, default: float = 0.0) -> float:
    try:
        return float(raw_value)
    except Exception:
        return default


def normalize_vault_classification(
    classification: Dict[str, Any],
    *,
    candidate_categories: Optional[List[str]] = None,
    fallback_primary: Optional[str] = None,
    fallback_confidence: Optional[float] = None,
    fallback_secondary: Optional[List[str]] = None,
) -> Dict[str, Any]:
    normalized = dict(classification or {})
    allowed_categories = {obsidian_tag(item) for item in (candidate_categories or []) if str(item).strip()}
    if not allowed_categories:
        allowed_categories = {obsidian_tag(item) for item in load_categories() if str(item).strip()}
    fallback_confidence_value = _coerce_float(fallback_confidence, 0.0)
    fallback_primary_tag = obsidian_tag(fallback_primary or "")
    primary = obsidian_tag(str(normalized.get("primary_label", "") or ""))
    secondary = _coerce_secondary_labels(normalized.get("secondary_labels", []))
    recovered_from_fallback = False
    recovered_from_secondary = False

    def is_allowed(label: str) -> bool:
        return not allowed_categories or label in allowed_categories

    if primary in {"", "unknown"} or not is_allowed(primary):
        if fallback_primary_tag and fallback_primary_tag not in {"unknown", "needs-review"} and is_allowed(fallback_primary_tag):
            normalized["primary_label"] = fallback_primary_tag
            primary = fallback_primary_tag
            recovered_from_fallback = True
        else:
            for candidate in secondary + list(fallback_secondary or []):
                candidate_tag = obsidian_tag(candidate)
                if candidate_tag in {"", "unknown", "needs-review"}:
                    continue
                if is_allowed(candidate_tag):
                    normalized["primary_label"] = candidate_tag
                    primary = candidate_tag
                    recovered_from_fallback = True
                    recovered_from_secondary = True
                    break
        if primary in {"", "unknown"} or not is_allowed(primary):
            if "needs-review" in allowed_categories:
                normalized["primary_label"] = "needs-review"
                primary = "needs-review"
                recovered_from_fallback = True
            else:
                normalized["primary_label"] = "unknown"
                primary = "unknown"

    secondary_tags: List[str] = []
    seen_secondary: set[str] = set()
    for item in secondary + list(fallback_secondary or []):
        tag = obsidian_tag(item)
        if tag in {"", "unknown", primary}:
            continue
        if allowed_categories and tag not in allowed_categories:
            continue
        if tag in seen_secondary:
            continue
        seen_secondary.add(tag)
        secondary_tags.append(tag)
    normalized["secondary_labels"] = secondary_tags

    confidence_raw = normalized.get("confidence")
    confidence = _coerce_float(confidence_raw, default=fallback_confidence_value)
    if confidence_raw is None or (isinstance(confidence_raw, str) and not confidence_raw.strip()):
        recovered_from_fallback = True
    else:
        try:
            float(confidence_raw)
        except Exception:
            recovered_from_fallback = True

    if recovered_from_fallback:
        if confidence <= 0:
            confidence = fallback_confidence_value if fallback_confidence_value > 0 else 0.55
        if not recovered_from_secondary:
            confidence = min(confidence, 0.69)
        normalized["confidence"] = round(confidence, 4)
        normalized.setdefault(
            "summary",
            "Classifier response was partially malformed, so the best available hybrid hint was used.",
        )
        normalized.setdefault(
            "reason",
            "Recovered a structured label from hybrid hints because the model response omitted required label fields.",
        )
        if not recovered_from_secondary:
            normalized["recommended_action"] = "review"
        normalized.setdefault("sensitive_flags", ["none"])
        normalized.setdefault("file_date_guess", "unknown")
        normalized.setdefault("language", "unknown")
    elif "confidence" in normalized:
        normalized["confidence"] = confidence

    if primary == "appeal":
        normalized["primary_label"] = "medical"
        if "appeals" not in normalized["secondary_labels"]:
            normalized["secondary_labels"] = ["appeals", *normalized["secondary_labels"]]

    return normalized


def classification_requires_review(classification: Dict[str, Any]) -> bool:
    primary = obsidian_tag(str((classification or {}).get("primary_label", "") or ""))
    confidence = _coerce_float((classification or {}).get("confidence"), 0.0)
    return primary in {"unknown", "needs-review"} or confidence < 0.70


def apply_text_message_override(
    classification: Dict[str, Any],
    *,
    source_path: str | Path | None,
    candidate_categories: Optional[List[str]] = None,
) -> Dict[str, Any]:
    if not is_text_message_source_path(source_path):
        return classification

    normalized = normalize_vault_classification(
        classification,
        candidate_categories=candidate_categories,
    )
    previous_primary = obsidian_tag(str(normalized.get("primary_label", "") or ""))
    merged_secondary: List[str] = []
    seen_secondary: set[str] = set()

    for item in [previous_primary, *(normalized.get("secondary_labels", []) or [])]:
        tag = obsidian_tag(str(item))
        if tag in {"", "unknown", "needs-review", "text-message"}:
            continue
        if tag in seen_secondary:
            continue
        seen_secondary.add(tag)
        merged_secondary.append(tag)

    candidate_categories_used: List[str] = []
    for item in ["text-message", *merged_secondary, *((normalized.get("candidate_categories_used", []) or []))]:
        tag = obsidian_tag(str(item))
        if tag and tag not in candidate_categories_used:
            candidate_categories_used.append(tag)

    return {
        **normalized,
        "primary_label": "text-message",
        "secondary_labels": merged_secondary,
        "candidate_categories_used": candidate_categories_used or ["text-message", "unknown", "needs-review"],
    }


def vault_category_parts(primary: str, secondary: List[Any]) -> List[str]:
    primary_tag = obsidian_tag(primary)
    secondary_tags = {obsidian_tag(item) for item in secondary}
    if primary_tag == "medical" and ({"appeal", "appeals"} & secondary_tags):
        return ["medical", "appeals"]
    return [primary_tag]


def vault_category_label(parts: List[str]) -> str:
    return " - ".join(parts)


def apply_codex_arbiter_if_enabled(
    *,
    enabled: bool,
    source_path: Path,
    markdown: str,
    local_classification: Dict[str, Any],
    heuristic_classification: Optional[Dict[str, Any]],
    hybrid_live_source: str,
) -> tuple[Dict[str, Any], str, Dict[str, Any]]:
    if not enabled:
        return local_classification, hybrid_live_source, {"status": "disabled", "applied": False}

    candidate_categories: List[str] = []
    for item in [
        *(local_classification.get("candidate_categories_used", []) or []),
        local_classification.get("primary_label", ""),
        *((local_classification.get("secondary_labels", []) or [])),
        str((heuristic_classification or {}).get("primary_label", "") or ""),
        "needs-review",
        "unknown",
    ]:
        tag = obsidian_tag(str(item))
        if tag and tag not in candidate_categories:
            candidate_categories.append(tag)

    classification, meta = run_codex_final_arbiter(
        source_path=source_path,
        markdown=markdown,
        local_classification=local_classification,
        candidate_categories=candidate_categories,
        command=RUNTIME_SETTINGS.codex_arbiter_command,
        timeout_seconds=RUNTIME_SETTINGS.codex_arbiter_timeout_seconds,
    )
    live_source = "codex-final-arbiter" if meta.get("applied") else hybrid_live_source
    return classification, live_source, meta


def yaml_list(items: List[Any]) -> str:
    if not items:
        return "[]"
    return "[" + ", ".join(json.dumps(str(x), ensure_ascii=False) for x in items) + "]"


def _sensitive_flags_from_labels(primary: str, secondary: List[str]) -> List[str]:
    labels = {obsidian_tag(primary), *(obsidian_tag(item) for item in secondary)}
    flags: List[str] = []
    if labels & {"medical", "appeal", "appeals", "benefits", "claim", "medical-receipt", "pharmacy", "prescription"}:
        flags.append("medical")
    if labels & {"receipt", "invoice", "financial", "tax", "insurance", "reimbursement-packet", "benefits", "claim"}:
        flags.append("financial")
    if labels & {"legal", "contract", "policy"}:
        flags.append("legal")
    if labels & {"identity-document"}:
        flags.append("identity")
    return flags or ["none"]


def build_reviewed_override_classification(
    *,
    reviewed_row: Dict[str, Any],
    source_path: Path,
    taxonomy_candidates: List[str],
) -> Dict[str, Any]:
    primary = obsidian_tag(
        str(
            reviewed_row.get("correct_label")
            or reviewed_row.get("primary_label")
            or reviewed_row.get("label")
            or "unknown"
        )
    )
    secondary: List[str] = []
    seen_secondary: set[str] = set()
    for item in reviewed_row.get("secondary_labels", []) or []:
        tag = obsidian_tag(str(item))
        if tag in {"", "unknown", primary} or tag in seen_secondary:
            continue
        seen_secondary.add(tag)
        secondary.append(tag)

    candidate_categories_used: List[str] = []
    for item in [primary, *secondary, *taxonomy_candidates]:
        tag = obsidian_tag(str(item))
        if tag and tag not in candidate_categories_used:
            candidate_categories_used.append(tag)

    review_status = str(reviewed_row.get("review_status") or "reviewed-feedback").strip() or "reviewed-feedback"
    summary = (
        str(reviewed_row.get("summary") or "").strip()
        or f"Used reviewed correction for {source_path.name}."
    )
    note = str(reviewed_row.get("note") or "").strip()
    reason = f"Used exact reviewed feedback for this source file from {review_status}."
    if note:
        reason = f"{reason} Evidence: {note}"

    return {
        "primary_label": primary,
        "secondary_labels": secondary,
        "confidence": 1.0,
        "summary": summary,
        "reason": reason,
        "sensitive_flags": _sensitive_flags_from_labels(primary, secondary),
        "recommended_action": "keep",
        "file_date_guess": "unknown",
        "language": "unknown",
        "candidate_categories_used": candidate_categories_used,
    }


def build_note_contract_metadata(
    *,
    source_path: Path,
    file_hash: str,
    attachment_link: str,
    attachment_mode: str,
    source_link: str = "",
    canonical_source_path: str | None = None,
    canonical_source_hash: str | None = None,
    last_seen_filename: str | None = None,
) -> Dict[str, str]:
    return {
        "canonical_source_path": canonical_source_path or str(source_path),
        "canonical_source_hash": canonical_source_hash or file_hash,
        "last_seen_filename": last_seen_filename or source_path.name,
        "attachment_mode": attachment_mode,
        "compatibility_attachment_path": attachment_link if attachment_mode == "copied-compatibility" else "",
        "source_link": source_link,
    }


def display_source_name(
    *,
    source_path: Path,
    canonical_source_path: str | None = None,
    last_seen_filename: str | None = None,
) -> str:
    for candidate in (last_seen_filename, canonical_source_path):
        if candidate:
            name = Path(candidate).name.strip()
            if name:
                return name
    return source_path.name


def build_canonical_source_link(canonical_source_path: str | None, display_name: str) -> str:
    if not canonical_source_path:
        return ""

    source_path = canonical_source_path.strip().replace("\\", "/")
    if not source_path:
        return ""

    cloud_vault_prefix = "/srv/cloud-vault/"
    if source_path.startswith(cloud_vault_prefix):
        relative_path = source_path[len(cloud_vault_prefix):]
        base_target = os.getenv(
            "CLASSIFIER_SOURCE_LINK_BASE_URL",
            r"\\192.168.50.86\cloud-vault",
        ).strip()
        if base_target.startswith("\\\\"):
            normalized_base = base_target.rstrip("\\/")
            target = normalized_base + "\\" + relative_path.replace("/", "\\")
        elif re.match(r"^[A-Za-z]:[/\\]", base_target):
            normalized_base = base_target.rstrip("\\/")
            target = normalized_base + "\\" + relative_path.replace("/", "\\")
        else:
            base_url = base_target.rstrip("/")
            target = f"{base_url}/{quote(relative_path, safe='/')}"
    elif source_path.startswith("/"):
        target = f"file://{quote(source_path, safe='/')}"
    elif re.match(r"^[A-Za-z]:/", source_path):
        target = f"file:///{quote(source_path, safe='/:')}"
    else:
        target = f"file://{quote(source_path, safe='/')}"

    label = display_name.replace("]", r"\]")
    return f"[{label}](<{target}>)"


def _build_vault_attachment_link(
    *,
    vault: Path,
    category_path: Path,
    visible_source_name: str,
    source_path: Path,
) -> tuple[str, str]:
    attachment_dir = vault / "90 Attachments" / category_path
    attachment_dir.mkdir(parents=True, exist_ok=True)
    copied = attachment_dir / build_attachment_filename(
        source_name=visible_source_name,
        existing_names={path.name for path in attachment_dir.iterdir() if path.is_file()},
    )
    if not copied.exists():
        shutil.copy2(source_path, copied)
    return f"[[{copied.relative_to(vault).as_posix()}]]", "copied-compatibility"


def _parse_note_frontmatter(text: str) -> dict[str, str]:
    lines = text.splitlines()
    if not lines or lines[0].strip() != "---":
        return {}

    values: dict[str, str] = {}
    for line in lines[1:]:
        stripped = line.strip()
        if stripped == "---":
            break
        if ":" not in line:
            continue
        key, raw_value = line.split(":", 1)
        key = key.strip()
        raw_value = raw_value.strip()
        try:
            parsed_value = json.loads(raw_value)
        except json.JSONDecodeError:
            parsed_value = raw_value.strip("'\"")
        if isinstance(parsed_value, str):
            values[key] = parsed_value
    return values


def _note_suffix_rank(name: str) -> tuple[int, int]:
    match = re.search(r" \((\d+)\)\.md$", name)
    if match is None:
        return (0, 0)
    return (1, int(match.group(1)))


def _has_legacy_hash_noise(name: str) -> bool:
    stem = Path(name).stem.lower()
    return bool(
        re.search(r" - [0-9a-f]{8,}$", stem)
        or re.match(r"^[0-9a-f]{16,}-", stem)
    )


def _extract_wikilink_target(value: str) -> PurePosixPath | None:
    raw_value = str(value or "").strip()
    match = re.fullmatch(r"\[\[([^\]]+)\]\]", raw_value)
    if match is None:
        return None
    inner = match.group(1).split("|", 1)[0].strip()
    if not inner:
        return None
    return PurePosixPath(inner)


def _existing_generated_note_matches(
    vault: Path,
    *,
    canonical_source_path: str,
    canonical_source_hash: str,
    last_seen_filename: str,
) -> list[tuple[int, Path, dict[str, str]]]:
    matches: list[tuple[int, Path, dict[str, str]]] = []
    for root_name in ("01 Classified", "02 Needs Review"):
        note_root = vault / root_name
        if not note_root.exists():
            continue
        for note_path in note_root.rglob("*.md"):
            metadata = _parse_note_frontmatter(
                note_path.read_text(encoding="utf-8", errors="replace")
            )
            if metadata.get("type") != "classified-document":
                continue
            rank: int | None = None
            if canonical_source_path and metadata.get("canonical_source_path") == canonical_source_path:
                rank = 0
            elif canonical_source_hash and metadata.get("canonical_source_hash") == canonical_source_hash:
                rank = 1
            elif last_seen_filename and metadata.get("last_seen_filename") == last_seen_filename:
                rank = 2
            if rank is not None:
                matches.append((rank, note_path, metadata))
    matches.sort(
        key=lambda item: (
            item[0],
            1 if _has_legacy_hash_noise(item[1].name) else 0,
            _note_suffix_rank(item[1].name),
            str(item[1]).lower(),
        )
    )
    return matches


def _resolve_note_attachment_path(vault: Path, metadata: dict[str, str], field_name: str) -> Path | None:
    target = _extract_wikilink_target(metadata.get(field_name, ""))
    if target is None:
        return None
    return (vault / Path(*target.parts)).resolve()


def _unlink_if_exists(path: Path | None) -> None:
    if path is None or not path.exists() or not path.is_file():
        return
    path.unlink()


def _remove_duplicate_generated_note(vault: Path, note_path: Path, metadata: dict[str, str]) -> None:
    _unlink_if_exists(_resolve_note_attachment_path(vault, metadata, "extracted_markdown"))
    note_path.unlink(missing_ok=True)


def build_summary_fallback(
    *,
    summary: str,
    source_path: Path,
    confidence: float,
    needs_review: bool,
) -> str:
    cleaned_summary = str(summary or "").strip()
    if cleaned_summary:
        return cleaned_summary
    if needs_review:
        return (
            f"Review needed for {source_path.name} because the classifier confidence "
            f"was {confidence:.2f} and the result still needs verification."
        )
    return f"Automatically classified {source_path.name} without a generated summary."


def build_reason_fallback(
    *,
    reason: str,
    needs_review: bool,
    primary_label: str,
) -> str:
    cleaned_reason = str(reason or "").strip()
    if cleaned_reason:
        return cleaned_reason
    if needs_review:
        return (
            "This file was routed to Needs Review because the classifier could not "
            "make a confident decision."
        )
    return f"No additional classifier reasoning was provided for {primary_label}."


def _is_ignored_input_folder_name(name: str) -> bool:
    cleaned = str(name or "").strip()
    return cleaned.startswith("_") or cleaned.startswith(".")


def _path_has_ignored_input_folder(path: Path, *, root: Path | None = None) -> bool:
    try:
        parts = path.relative_to(root).parts if root is not None else path.parts
    except Exception:
        parts = path.parts
    directory_parts = parts[:-1] if path.suffix else parts
    return any(_is_ignored_input_folder_name(part) for part in directory_parts)


def iter_input_files(path: Path) -> Iterable[Path]:
    if path.is_file():
        if path.suffix.lower() in SUPPORTED_EXTENSIONS and not _path_has_ignored_input_folder(path):
            yield path
        return

    for current_root, dir_names, file_names in os.walk(path):
        dir_names[:] = [
            dir_name
            for dir_name in dir_names
            if not _is_ignored_input_folder_name(dir_name)
        ]
        current_root_path = Path(current_root)
        for file_name in file_names:
            candidate = current_root_path / file_name
            if candidate.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue
            if _path_has_ignored_input_folder(candidate, root=path):
                continue
            yield candidate

def ensure_vault(vault: Path) -> None:
    dirs = [
        vault,
        vault / ".obsidian",
        vault / "00 Inbox",
        vault / "01 Classified",
        vault / "02 Needs Review",
        vault / "90 Attachments",
        vault / "_system",
        vault / "_system" / "classifications",
        vault / "_system" / "extracted-markdown",
        vault / "_system" / "templates",
    ]

    for d in dirs:
        d.mkdir(parents=True, exist_ok=True)

    app_json = vault / ".obsidian" / "app.json"
    if not app_json.exists():
        app_json.write_text(json.dumps({
            "alwaysUpdateLinks": True,
            "newFileLocation": "current",
            "attachmentFolderPath": "90 Attachments"
        }, indent=2), encoding="utf-8")

    community_plugins = vault / ".obsidian" / "community-plugins.json"
    if not community_plugins.exists():
        community_plugins.write_text("[]\n", encoding="utf-8")

    home = vault / "Home.md"
    if not home.exists():
        home.write_text("""---
type: vault-home
system: local-document-classifier
---

# Local Document Classifier

Generated Obsidian vault.
""", encoding="utf-8")

def parse_plain_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")

def convert_legacy_office(path: Path, work_dir: Path) -> Optional[Path]:
    ext = path.suffix.lower()
    target_ext = {
        ".doc": "docx",
        ".xls": "xlsx",
        ".ppt": "pptx",
    }.get(ext)

    if not target_ext:
        return None

    work_dir.mkdir(parents=True, exist_ok=True)

    cmd = [
        "soffice",
        "--headless",
        "--convert-to",
        target_ext,
        "--outdir",
        str(work_dir),
        str(path),
    ]

    try:
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=180)
    except Exception:
        return None

    converted = work_dir / f"{path.stem}.{target_ext}"
    return converted if converted.exists() else None

def parse_with_docling(path: Path) -> str:
    from docling.document_converter import DocumentConverter

    converter = DocumentConverter()
    result = converter.convert(str(path))
    return result.document.export_to_markdown()

def parse_docx_fast(path: Path) -> tuple[str, str]:
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: List[str] = []

    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")

    root = ET.fromstring(xml)
    for paragraph in root.findall(".//w:p", namespace):
        parts = []
        for node in paragraph.findall(".//w:t", namespace):
            if node.text:
                parts.append(node.text)
        if parts:
            paragraphs.append("".join(parts))

    text = "\n".join(paragraphs).strip()
    if not text:
        raise RuntimeError(f"No text extracted from DOCX: {path}")
    return text, "docx-xml"

def parse_pdf_fast(path: Path) -> tuple[str, str]:
    cmd = [
        "pdftotext",
        "-layout",
        "-nopgbrk",
        str(path),
        "-",
    ]
    proc = subprocess.run(
        cmd,
        check=False,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=180,
    )
    text = (proc.stdout or "").strip()
    if proc.returncode != 0 or len(text) < 80:
        raise RuntimeError(f"pdftotext did not produce usable text for {path}")
    return text, "pdftotext"


def parse_pdf_with_ocr_fallback(path: Path) -> tuple[str, str, Dict[str, Any]]:
    result = extract_pdf_text_with_metadata(path.read_bytes(), source_name=path.name)
    text = str(result.get("text", "") or "").strip()
    parser_name = str(result.get("parser", "") or "pdf-ocr")
    if not text:
        raise RuntimeError(f"No text extracted from PDF after OCR fallback: {path}")
    return text, parser_name, {
        "ocr_engine": str(result.get("ocr_engine", "") or ""),
        "ocr_quality": str(result.get("quality", "empty") or "empty"),
        "ocr_char_count": sum(1 for char in text if char.isalnum()),
        "extraction_quality": str(result.get("quality", "empty") or "empty"),
    }

def build_fast_document_classification(
    primary: str,
    secondary: List[str],
    confidence: float,
    summary: str,
    reason: str,
    categories: List[str],
) -> Dict[str, Any]:
    allowed = set(categories)
    return {
        "primary_label": primary if primary in allowed else "unknown",
        "secondary_labels": [label for label in secondary if label in allowed and label != primary],
        "confidence": confidence,
        "summary": summary,
        "reason": reason,
        "sensitive_flags": [
            label for label in ["legal", "financial", "medical", "insurance", "tax"] if label in {primary, *secondary}
        ] or ["none"],
        "recommended_action": "review" if primary == "legal" else "keep",
        "file_date_guess": "unknown",
        "language": "English",
        "candidate_categories_used": [
            label for label in ["legal", "contract", "policy", "work", "technical", "report", "unknown", "needs-review"] if label in allowed
        ],
    }

def classify_document_fast(
    source_path: Path,
    markdown: str,
    categories: List[str],
    rules: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    text = f"{source_path.name}\n{markdown}".lower()
    rules = rules or load_heuristic_rules()
    document_rules = rules.get("document_fast_path", {})

    def score(keywords: List[str]) -> int:
        return sum(1 for keyword in keywords if keyword in text)

    legal_terms = [
        "agreement", "scope of services", "term", "confidentiality", "payment",
        "limitation of liability", "governing law", "termination", "party", "parties",
        "vendor", "service agreement", "contract",
    ]
    technical_terms = [
        "incident", "severity", "environment", "timeline", "packet loss",
        "route", "bgp", "rollback", "root cause", "corrective actions",
        "network", "staging", "monitoring", "runbook", "policy",
    ]
    report_terms = [
        "incident overview", "incident metadata", "timeline", "root cause",
        "corrective actions", "status", "resolved", "report",
    ]
    policy_terms = [
        "policy", "procedure", "validation", "controls", "runbook",
    ]

    legal_score = score(legal_terms)
    technical_score = score(technical_terms)
    report_score = score(report_terms)
    policy_score = score(policy_terms)

    if legal_score >= int(document_rules.get("legal_agreement_min_score", 6)) and legal_score >= technical_score + 2:
        return build_fast_document_classification(
            primary="legal",
            secondary=["contract", "work"] + (["policy"] if policy_score > 0 else []),
            confidence=0.96,
            summary="Text-extractable service agreement with clear contract sections and legal terms.",
            reason=(
                "Fast document path used because the extracted text strongly matches a legal/service agreement "
                "with scope, term, confidentiality, payment, liability, and governing-law sections."
            ),
            categories=categories,
        )

    if (
        technical_score >= int(document_rules.get("technical_incident_min_score", 6))
        and report_score >= int(document_rules.get("technical_report_min_score", 4))
    ):
        return build_fast_document_classification(
            primary="report",
            secondary=["work", "technical", "policy"],
            confidence=0.95,
            summary="Technical incident report with timeline, root cause, and corrective actions.",
            reason=(
                "Fast document path used because the extracted text clearly describes a network incident report "
                "with severity, timeline, root cause, rollback, and corrective actions."
            ),
            categories=categories,
        )

    return None

def parse_spreadsheet_fast(
    path: Path,
    max_sheets: int = 4,
    max_rows: int = 8,
    max_cols: int = 8,
    max_cell_chars: int = 80,
) -> tuple[str, str, Dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet_summaries: List[Dict[str, Any]] = []
    non_empty_preview_cells = 0

    for worksheet in workbook.worksheets[:max_sheets]:
        title_lower = worksheet.title.lower()
        is_metadata_sheet = any(token in title_lower for token in ["expectation", "fixture", "metadata"])
        preview_rows: List[List[str]] = []

        if not is_metadata_sheet:
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_index > max_rows:
                    break

                values: List[str] = []
                has_value = False

                for cell in row[:max_cols]:
                    value = "" if cell is None else str(cell).strip().replace("\n", " ")
                    value = value[:max_cell_chars]
                    if value:
                        has_value = True
                        non_empty_preview_cells += 1
                    values.append(value)

                if has_value:
                    preview_rows.append(values)

        sheet_summaries.append(
            {
                "title": worksheet.title,
                "preview_rows": preview_rows,
                "is_metadata_sheet": is_metadata_sheet,
            }
        )

    lines = [
        "# Spreadsheet Summary",
        f"Workbook: {path.name}",
        f"Sheet count: {len(workbook.sheetnames)}",
    ]

    for sheet in sheet_summaries:
        lines.append("")
        lines.append(f"## Sheet: {sheet['title']}")

        preview_rows = sheet["preview_rows"]
        if sheet.get("is_metadata_sheet"):
            lines.append("- Metadata-style sheet omitted from classification preview.")
            continue

        if not preview_rows:
            lines.append("- No preview rows with visible values.")
            continue

        for row in preview_rows[:6]:
            cells = [cell for cell in row if cell]
            if cells:
                lines.append(f"- {' | '.join(cells)}")

    markdown = "\n".join(lines).strip()
    metadata = {
        "sheet_count": len(workbook.sheetnames),
        "sheet_names": workbook.sheetnames[:max_sheets],
        "preview_sheet_count": len(sheet_summaries),
        "non_empty_preview_cells": non_empty_preview_cells,
    }
    return markdown, "spreadsheet-openpyxl", metadata

def classify_spreadsheet_fast(
    source_path: Path,
    categories: List[str],
    markdown: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
) -> tuple[str, Dict[str, Any], Dict[str, Any]]:
    if markdown is None or metadata is None:
        markdown, parser_name, metadata = parse_spreadsheet_fast(source_path)
        metadata["parser"] = parser_name
    else:
        metadata = dict(metadata)
        metadata.setdefault("parser", "spreadsheet-openpyxl")

    text = f"{source_path.name}\n{markdown}".lower()
    available = set(categories)

    def score(*keywords: str) -> int:
        return sum(1 for keyword in keywords if keyword in text)

    domain_scores = {
        "financial": score(
            "budget", "forecast", "variance", "quarter", "fy forecast",
            "actual", "cost", "expense", "revenue", "payment", "department"
        ),
        "tax": score("tax", "irs", "1099", "w-2", "withholding", "deduction"),
        "legal": score("agreement", "contract", "terms", "effective date", "vendor"),
        "medical": score("patient", "medical", "diagnosis", "provider", "prescription"),
        "insurance": score("insurance", "claim", "coverage", "premium", "policy"),
        "technical": score("incident", "network", "server", "uptime", "ticket", "configuration"),
    }

    primary = "spreadsheet" if "spreadsheet" in available else None

    if domain_scores["tax"] >= 2 and "tax" in available:
        primary = "tax"
    elif domain_scores["medical"] >= 2 and "medical" in available:
        primary = "medical"
    elif domain_scores["insurance"] >= 2 and "insurance" in available:
        primary = "insurance"
    elif domain_scores["legal"] >= 2 and "legal" in available:
        primary = "legal"
    elif domain_scores["technical"] >= 3 and "technical" in available:
        primary = "technical"
    elif primary is None and domain_scores["financial"] >= 2 and "financial" in available:
        primary = "financial"

    if primary is None:
        for fallback in ["financial", "work", "report", "unknown", "needs-review"]:
            if fallback in available:
                primary = fallback
                break
        primary = primary or "unknown"

    secondary: List[str] = []
    for label in ["spreadsheet", "financial", "work", "report", "tax", "legal", "medical", "insurance", "technical"]:
        if label != primary and label in available:
            if label == "financial" and domain_scores["financial"] == 0:
                continue
            if label == "technical" and domain_scores["technical"] < 2:
                continue
            if label in {"tax", "legal", "medical", "insurance"} and domain_scores[label] == 0:
                continue
            secondary.append(label)

    secondary = secondary[:4]

    sensitive_flags = [
        label
        for label in ["financial", "tax", "legal", "medical", "insurance"]
        if domain_scores.get(label, 0) > 0
    ] or ["none"]

    confidence = 0.98 if primary == "spreadsheet" else 0.94
    sheet_names = ", ".join(metadata.get("sheet_names", [])[:3]) or "workbook preview"
    reason_parts = [
        "Fast spreadsheet path used to avoid slow OCR/LLM processing.",
        f"Detected workbook structure across {metadata.get('sheet_count', 0)} sheet(s): {sheet_names}.",
    ]
    if domain_scores["financial"] > 0:
        reason_parts.append("Budget and forecast terms indicate financial spreadsheet content.")

    classification = {
        "primary_label": primary,
        "secondary_labels": secondary,
        "confidence": confidence,
        "summary": (
            f"Spreadsheet workbook preview with {metadata.get('sheet_count', 0)} sheet(s), "
            f"including {sheet_names}."
        ),
        "reason": " ".join(reason_parts),
        "sensitive_flags": sensitive_flags,
        "recommended_action": "keep",
        "file_date_guess": "unknown",
        "language": "English",
        "candidate_categories_used": [
            label
            for label in ["spreadsheet", "financial", "work", "report", "tax", "legal", "medical", "insurance", "technical", "unknown", "needs-review"]
            if label in available
        ],
    }
    return markdown, classification, metadata


def resolve_hybrid_document_decision(
    source_path: Path,
    markdown: str,
    parser_name: str,
    categories: List[str],
    heuristic_result: Optional[Dict[str, Any]],
    ollama_url: str,
    model: str,
    max_chars: int,
    *,
    reviewed_source_path: str | Path | None = None,
    lightgbm_model_path: Optional[Path] = None,
    gating_config: Optional[Dict[str, Any]] = None,
    timing: Optional[Dict[str, Any]] = None,
    extraction_metadata: Optional[Dict[str, Any]] = None,
) -> tuple[Dict[str, Any], Dict[str, Any]]:
    extraction_metadata = dict(extraction_metadata or {})
    reviewed_source_path_text = str(reviewed_source_path or source_path).strip()
    retrieval_metadata = build_retrieval_metadata(
        source_path=source_path,
        text=markdown,
        classification=heuristic_result,
    )
    taxonomy_candidates = select_candidate_categories(
        all_categories=categories,
        filename=source_path.name,
        extension=source_path.suffix.lower(),
        content=markdown,
        is_image=False,
    )

    reviewed_override = find_reviewed_label_override(
        source_path=reviewed_source_path_text,
        filename=source_path.name,
        limit=1500,
    )
    override_source_path = str((reviewed_override or {}).get("source_path", "") or "").strip()
    if reviewed_override and override_source_path == reviewed_source_path_text:
        decision = {
            "use_inline_llm": False,
            "live_source": "manual-correction-override",
            "selected_primary_hint": str(
                reviewed_override.get("correct_label")
                or reviewed_override.get("primary_label")
                or reviewed_override.get("label")
                or "unknown"
            ),
            "decision_reason": "exact-source-reviewed-feedback",
            "candidate_count": len(taxonomy_candidates),
        }
        if timing is not None:
            timing["hybrid_live_source"] = decision["live_source"]
            timing["hybrid_decision_reason"] = decision["decision_reason"]
        classification = build_reviewed_override_classification(
            reviewed_row=reviewed_override,
            source_path=source_path,
            taxonomy_candidates=taxonomy_candidates,
        )
        hybrid_meta = {
            "taxonomy_candidates": taxonomy_candidates,
            "lightgbm": None,
            "decision": decision,
            "retrieval": retrieval_metadata,
            "extraction": extraction_metadata,
            "reviewed_override": reviewed_override,
        }
        return classification, hybrid_meta

    lightgbm_result: Optional[Dict[str, Any]] = None
    model_path = lightgbm_model_path or LIGHTGBM_MODEL_PATH
    try:
        lightgbm_result = predict_lightgbm_result(
            payload={
                "filename": source_path.name,
                "extension": source_path.suffix.lower(),
                "parser": parser_name,
                "text_preview": markdown,
                "heuristic_primary": (heuristic_result or {}).get("primary_label", "unknown"),
                "taxonomy_candidates": taxonomy_candidates,
                **extraction_metadata,
                **retrieval_metadata,
            },
            model_path=model_path,
        )
    except Exception:
        lightgbm_result = None

    if heuristic_result is None:
        decision = {
            "use_inline_llm": True,
            "live_source": "inline-llm",
            "selected_primary_hint": (lightgbm_result or {}).get("top_label", "unknown"),
            "decision_reason": "no-heuristic-fast-path",
        }
    elif lightgbm_result is None:
        decision = {
            "use_inline_llm": False,
            "live_source": "heuristic-fast-path",
            "selected_primary_hint": heuristic_result.get("primary_label", "unknown"),
            "decision_reason": "heuristic-without-lightgbm",
        }
    else:
        decision = choose_live_decision(
            heuristic_result=heuristic_result,
            lightgbm_result=lightgbm_result,
            gating_config=gating_config or load_hybrid_gating_config(),
            candidate_categories=taxonomy_candidates,
        )
        forced_keys = set(load_heuristic_rules().get("force_inline_llm_for", []) or [])
        force_key = f"{parser_name}|{heuristic_result.get('primary_label', 'unknown')}"
        if force_key in forced_keys:
            decision = {
                **decision,
                "use_inline_llm": True,
                "live_source": "inline-llm",
                "decision_reason": "forced-inline-from-disagreement-config",
            }
    fast_path_preview = apply_text_message_override(
        heuristic_result or {},
        source_path=reviewed_source_path_text,
        candidate_categories=taxonomy_candidates,
    )
    if not decision["use_inline_llm"] and classification_requires_review(fast_path_preview):
        decision = {
            **decision,
            "use_inline_llm": True,
            "live_source": "inline-llm",
            "decision_reason": "review-bound-inline-llm",
        }

    if timing is not None:
        timing["hybrid_live_source"] = decision.get("live_source")
        timing["hybrid_decision_reason"] = decision.get("decision_reason")
        if lightgbm_result:
            timing["lightgbm_top_label"] = lightgbm_result.get("top_label")
            timing["lightgbm_top_probability"] = lightgbm_result.get("top_probability")
            timing["lightgbm_needs_llm_probability"] = lightgbm_result.get("needs_llm_probability")
            timing["lightgbm_disagreement_risk"] = lightgbm_result.get("disagreement_risk")

    if decision["use_inline_llm"]:
        classification = classify_markdown(
            markdown=markdown,
            source_path=source_path,
            reviewed_source_path=reviewed_source_path_text,
            categories=categories,
            ollama_url=ollama_url,
            model=model,
            max_chars=max_chars,
            timing=timing,
            heuristic_hints=heuristic_result,
        )
    else:
        classification = heuristic_result or {
            "primary_label": "unknown",
            "secondary_labels": [],
            "confidence": 0.0,
            "summary": "No classification available.",
            "reason": "No heuristic result and no inline model was used.",
            "sensitive_flags": ["none"],
            "recommended_action": "review",
            "file_date_guess": "unknown",
            "language": "unknown",
            "candidate_categories_used": taxonomy_candidates,
        }
    classification = apply_text_message_override(
        classification,
        source_path=reviewed_source_path_text,
        candidate_categories=taxonomy_candidates,
    )

    hybrid_meta = {
        "taxonomy_candidates": taxonomy_candidates,
        "lightgbm": lightgbm_result,
        "decision": decision,
        "retrieval": retrieval_metadata,
        "extraction": extraction_metadata,
    }
    return classification, hybrid_meta


def should_enqueue_shadow_job(gating_config: Dict[str, Any], live_source: str) -> bool:
    mode = str(gating_config.get("shadow_mode", "all"))
    if mode == "off":
        return False
    if mode == "all":
        return True
    if mode == "fast-only":
        return live_source == "heuristic-fast-path"
    return True


def process_shadow_queue_command(
    categories: List[str],
    ollama_url: str,
    model: str,
    vision_model: str,
    max_chars: int,
) -> Dict[str, Any]:
    gating_config = load_hybrid_gating_config()

    def shadow_classifier(job: Dict[str, Any]) -> Dict[str, Any]:
        mode = str(job.get("mode", "document"))
        if mode == "image":
            source_path = Path(str(job.get("source_path", "")))
            classification, _, _ = classify_image(
                source_path=source_path,
                reviewed_source_path=str(job.get("source_path", "") or source_path),
                categories=categories,
                ollama_url=ollama_url,
                model=model,
                vision_model=vision_model,
                max_chars=max_chars,
            )
            return normalize_vault_classification(
                classification,
                candidate_categories=list(job.get("taxonomy_candidates", []) or classification.get("candidate_categories_used", []) or []),
                fallback_primary=str(((job.get("lightgbm_result") or {}).get("top_label") or "")),
                fallback_confidence=float(((job.get("lightgbm_result") or {}).get("top_probability") or 0.0)),
                fallback_secondary=[str(((job.get("heuristic_result") or {}).get("primary_label") or ""))],
            )

        classification = classify_markdown(
            markdown=str(job.get("markdown", "")),
            source_path=Path(str(job.get("filename", "shadow-document"))),
            reviewed_source_path=str(job.get("source_path", "") or job.get("filename", "shadow-document")),
            categories=categories,
            ollama_url=ollama_url,
            model=model,
            max_chars=max_chars,
            heuristic_hints=job.get("heuristic_result") or {},
        )
        return normalize_vault_classification(
            classification,
            candidate_categories=list(job.get("taxonomy_candidates", []) or classification.get("candidate_categories_used", []) or []),
            fallback_primary=str(
                ((job.get("lightgbm_result") or {}).get("top_label"))
                or ((job.get("heuristic_result") or {}).get("primary_label"))
                or ""
            ),
            fallback_confidence=float(((job.get("lightgbm_result") or {}).get("top_probability") or 0.0)),
            fallback_secondary=[str(((job.get("heuristic_result") or {}).get("primary_label") or ""))],
        )

    return run_autonomous_shadow_cycle(
        shadow_classifier=shadow_classifier,
        gating_config=gating_config,
    )

def parse_document(path: Path, work_dir: Path) -> tuple[str, str, Dict[str, Any]]:
    ext = path.suffix.lower()

    if ext in SPREADSHEET_EXTENSIONS:
        markdown, parser_name, _ = parse_spreadsheet_fast(path)
        return markdown, parser_name, {
            "ocr_engine": "",
            "ocr_quality": "",
            "ocr_char_count": 0,
            "extraction_quality": extraction_quality_for_text(markdown),
        }

    if ext in DOCX_EXTENSIONS:
        try:
            markdown, parser_name = parse_docx_fast(path)
            return markdown, parser_name, {
                "ocr_engine": "",
                "ocr_quality": "",
                "ocr_char_count": 0,
                "extraction_quality": extraction_quality_for_text(markdown),
            }
        except Exception:
            pass

    if ext in PDF_EXTENSIONS:
        try:
            markdown, parser_name = parse_pdf_fast(path)
            return markdown, parser_name, {
                "ocr_engine": "",
                "ocr_quality": "",
                "ocr_char_count": 0,
                "extraction_quality": extraction_quality_for_text(markdown),
            }
        except Exception:
            return parse_pdf_with_ocr_fallback(path)

    if ext in PLAIN_EXTENSIONS:
        markdown = parse_plain_text(path)
        return markdown, "plain-text", {
            "ocr_engine": "",
            "ocr_quality": "",
            "ocr_char_count": 0,
            "extraction_quality": extraction_quality_for_text(markdown),
        }

    try:
        markdown = parse_with_docling(path)
        return markdown, "docling", {
            "ocr_engine": "",
            "ocr_quality": "",
            "ocr_char_count": 0,
            "extraction_quality": extraction_quality_for_text(markdown),
        }
    except Exception as first_error:
        converted = convert_legacy_office(path, work_dir)
        if converted:
            try:
                markdown = parse_with_docling(converted)
                return markdown, "docling-converted", {
                    "ocr_engine": "",
                    "ocr_quality": "",
                    "ocr_char_count": 0,
                    "extraction_quality": extraction_quality_for_text(markdown),
                }
            except Exception:
                pass

        if ext in {".html", ".htm"}:
            markdown = parse_plain_text(path)
            return markdown, "html-plain", {
                "ocr_engine": "",
                "ocr_quality": "",
                "ocr_char_count": 0,
                "extraction_quality": extraction_quality_for_text(markdown),
            }

        raise RuntimeError(f"Document parsing failed for {path}: {first_error}") from first_error

def extract_json(text: str) -> Dict[str, Any]:
    text = text.strip()
    decoder = JSONDecoder()

    if text.startswith("```"):
        fenced_lines = [line for line in text.splitlines() if not line.strip().startswith("```")]
        text = "\n".join(fenced_lines).strip()

    try:
        return json.loads(text)
    except Exception:
        pass

    match = re.search(r"\{.*\}", text, flags=re.DOTALL)
    candidates: List[str] = []
    if match:
        candidates.append(match.group(0))

    for candidate in [text, *candidates]:
        for start_index, char in enumerate(candidate):
            if char != "{":
                continue
            try:
                payload, _ = decoder.raw_decode(candidate[start_index:])
            except Exception:
                continue
            if isinstance(payload, dict):
                return payload

    raise ValueError(f"Could not parse JSON from model response: {text[:500]}")

def ollama_chat(
    ollama_url: str,
    model: str,
    messages: List[Dict[str, Any]],
    json_mode: bool = True,
    timeout: int = 600,
) -> str:
    payload: Dict[str, Any] = {
        "model": model,
        "messages": messages,
        "stream": False,
        "options": {
            "temperature": 0
        }
    }

    if json_mode:
        payload["format"] = "json"

    response = requests.post(
        f"{ollama_url.rstrip('/')}/api/chat",
        json=payload,
        timeout=timeout,
    )
    response.raise_for_status()
    data = response.json()
    return data["message"]["content"]

def classify_markdown(
    markdown: str,
    source_path: Path,
    categories: List[str],
    ollama_url: str,
    model: str,
    max_chars: int,
    *,
    reviewed_source_path: str | Path | None = None,
    timing: Optional[Dict[str, Any]] = None,
    heuristic_hints: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    ext = source_path.suffix.lower()
    clipped = markdown[:max_chars]

    candidate_categories = select_candidate_categories(
        all_categories=categories,
        filename=source_path.name,
        extension=ext,
        content=clipped,
        is_image=False,
    )

    examples = load_relevant_examples(
        filename=source_path.name,
        extension=ext,
        content=clipped,
        is_image=False,
        source_path=reviewed_source_path or source_path,
    )

    prompt = f"""
You are a private local document classifier.

Classify this file using only the filename, extracted document content, allowed categories, and prior correction examples.
Do not invent names, dates, dollar amounts, account numbers, or legal/medical facts.
If the content is ambiguous, lower confidence and use "unknown" or "needs-review".

Important:
- Choose primary_label only from Allowed categories.
- Use secondary_labels only from Allowed categories.
- A reimbursement packet with attached receipts may be "reimbursement-packet" or "receipt" depending on the dominant content.
- Return strict JSON only.
- Never return markdown fences, numbered lists, or explanatory prose outside the JSON object.
- If uncertain, still return the required JSON schema with primary_label set to "unknown" or "needs-review".

Allowed categories:
{json.dumps(candidate_categories, indent=2)}

Prior correction examples:
{format_examples_for_prompt(examples)}

Heuristic hints:
{json.dumps(heuristic_hints or {}, indent=2)}

Return exactly this JSON shape:
{{
  "primary_label": "one allowed category",
  "secondary_labels": ["zero or more allowed categories"],
  "confidence": 0.0,
  "summary": "short factual summary",
  "reason": "why this label was chosen",
  "sensitive_flags": ["pii", "medical", "financial", "legal", "identity", "none"],
  "recommended_action": "keep, review, archive, reimburse, delete, or unknown",
  "file_date_guess": "YYYY-MM-DD, YYYY-MM, YYYY, or unknown",
  "language": "detected language or unknown"
}}

Filename:
{source_path.name}

Extracted content:
{clipped}
""".strip()

    if timing is not None:
        timing["candidate_category_count"] = len(candidate_categories)
        timing["markdown_chars"] = len(markdown)
        timing["clipped_markdown_chars"] = len(clipped)
        timing["model"] = model

    model_started_at = time.perf_counter()
    content = ollama_chat(
        ollama_url=ollama_url,
        model=model,
        messages=[{"role": "user", "content": prompt}],
        json_mode=True,
    )
    if timing is not None:
        timing["model_ms"] = elapsed_ms(model_started_at)

    result = extract_json(content)
    result["candidate_categories_used"] = candidate_categories
    return result

def classify_image_vision(
    source_path: Path,
    categories: List[str],
    ollama_url: str,
    vision_model: str,
    timing: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    ext = source_path.suffix.lower()

    candidate_categories = select_candidate_categories(
        all_categories=categories,
        filename=source_path.name,
        extension=ext,
        content="",
        is_image=True,
    )
    candidate_categories = image_safe_categories(candidate_categories, source_path.name)

    examples = load_relevant_examples(
        filename=source_path.name,
        extension=ext,
        content="",
        is_image=True,
        source_path=source_path,
    )

    image_b64 = base64.b64encode(source_path.read_bytes()).decode("utf-8")

    prompt = f"""
You are a private local image classifier.

Look at the image and classify it using the allowed categories.

Critical rule:
If it is not a document, do not force it into receipt, legal, medical, insurance, tax, financial, technical, or marketing categories.

For environment/reference images, prefer labels such as reference-image, concept-art, environment-art, game-reference, architecture, industrial, sci-fi, snow-ice, frozen-environment, post-apocalyptic, waystation, facility, building, interior, exterior, machinery, artwork, photo, image-only, or unknown as appropriate.

Describe visible content only.
Do not invent unreadable text.
Return strict JSON only.
- Never return markdown fences, numbered lists, or explanatory prose outside the JSON object.
- If uncertain, still return the required JSON schema with primary_label set to "unknown" or "needs-review".

Allowed categories:
{json.dumps(candidate_categories, indent=2)}

Prior correction examples:
{format_examples_for_prompt(examples)}

Return exactly this JSON shape:
{{
  "primary_label": "one allowed category",
  "secondary_labels": ["zero or more allowed categories"],
  "confidence": 0.0,
  "summary": "short factual summary of visible content",
  "reason": "why this label was chosen",
  "sensitive_flags": ["pii", "medical", "financial", "legal", "identity", "none"],
  "recommended_action": "keep, review, archive, reimburse, delete, or unknown",
  "file_date_guess": "YYYY-MM-DD, YYYY-MM, YYYY, or unknown",
  "language": "detected language or unknown"
}}

Filename:
{source_path.name}
""".strip()

    if timing is not None:
        timing["candidate_category_count"] = len(candidate_categories)
        timing["image_bytes"] = source_path.stat().st_size
        timing["model"] = vision_model

    model_started_at = time.perf_counter()
    content = ollama_chat(
        ollama_url=ollama_url,
        model=vision_model,
        messages=[{
            "role": "user",
            "content": prompt,
            "images": [image_b64],
        }],
        json_mode=True,
        timeout=900,
    )
    if timing is not None:
        timing["model_ms"] = elapsed_ms(model_started_at)

    result = extract_json(content)
    result["candidate_categories_used"] = candidate_categories
    return normalize_image_classification_result(result)


def classify_image(
    source_path: Path,
    categories: List[str],
    ollama_url: str,
    model: str,
    vision_model: str,
    max_chars: int,
    *,
    reviewed_source_path: str | Path | None = None,
    gating_config: Optional[Dict[str, Any]] = None,
    heuristic_rules: Optional[Dict[str, Any]] = None,
    timing: Optional[Dict[str, Any]] = None,
) -> tuple[Dict[str, Any], Optional[Dict[str, Any]], str]:
    mime_type = mimetypes.guess_type(source_path.name)[0] or "image/png"
    ocr_evidence = extract_image_text_with_metadata(
        path=str(source_path),
        mime_type=mime_type,
        payload=source_path.read_bytes(),
    )
    ocr_text = str(ocr_evidence.get("text", "") or "")
    ocr_engine = str(ocr_evidence.get("engine", "") or "")
    ocr_quality = str(ocr_evidence.get("quality", "empty") or "empty")

    if timing is not None:
        timing["ocr_engine"] = ocr_engine
        timing["ocr_quality"] = ocr_quality
        timing["ocr_chars"] = int(ocr_evidence.get("char_count", 0) or 0)
        timing["extraction_quality"] = ocr_quality

    if int(ocr_evidence.get("char_count", 0) or 0) >= IMAGE_OCR_MIN_CHARS:
        parser_name = f"image-ocr-{ocr_engine or 'unknown'}"
        extraction_metadata = {
            "ocr_engine": ocr_engine,
            "ocr_quality": ocr_quality,
            "ocr_char_count": int(ocr_evidence.get("char_count", 0) or 0),
            "extraction_quality": ocr_quality,
        }
        heuristic_classification = classify_document_fast(
            source_path=source_path,
            markdown=ocr_text,
            categories=categories,
            rules=heuristic_rules,
        )
        classification, hybrid_meta = resolve_hybrid_document_decision(
            source_path=source_path,
            reviewed_source_path=reviewed_source_path,
            markdown=ocr_text,
            parser_name=parser_name,
            categories=categories,
            heuristic_result=heuristic_classification,
            ollama_url=ollama_url,
            model=model,
            max_chars=max_chars,
            gating_config=gating_config,
            timing=timing,
            extraction_metadata=extraction_metadata,
        )
        classification["ocr_engine"] = ocr_engine
        classification["ocr_quality"] = ocr_quality
        classification["ocr_char_count"] = int(ocr_evidence.get("char_count", 0) or 0)
        classification["extraction_quality"] = ocr_quality
        return classification, hybrid_meta, ocr_text

    classification = classify_image_vision(
        source_path=source_path,
        categories=categories,
        ollama_url=ollama_url,
        vision_model=vision_model,
        timing=timing,
    )
    classification = apply_text_message_override(
        classification,
        source_path=reviewed_source_path or source_path,
        candidate_categories=list(classification.get("candidate_categories_used", []) or []),
    )
    classification["ocr_engine"] = ocr_engine
    classification["ocr_quality"] = ocr_quality
    classification["ocr_char_count"] = int(ocr_evidence.get("char_count", 0) or 0)
    classification["extraction_quality"] = ocr_quality
    return classification, None, ""

def write_obsidian_note(
    vault: Path,
    source_path: Path,
    file_hash: str,
    markdown: Optional[str],
    classification: Dict[str, Any],
    attach_originals: bool,
    canonical_source_path: str | None = None,
    canonical_source_hash: str | None = None,
    last_seen_filename: str | None = None,
    source_parser: str | None = None,
    heuristic_primary_hint: str | None = None,
    hybrid_live_source: str | None = None,
) -> Path:
    # --- image-normalize-before-note BEGIN ---
    try:
        if source_path.suffix.lower() in IMAGE_EXTENSIONS:
            from .category_manager import normalize_image_classification_result
            classification = normalize_image_classification_result(classification)
    except Exception:
        pass
    # --- image-normalize-before-note END ---

    hybrid_fallback = {
        "candidate_categories": classification.get("candidate_categories_used", []) or [],
        "fallback_primary": classification.get("primary_label"),
        "fallback_confidence": classification.get("confidence"),
        "fallback_secondary": classification.get("secondary_labels", []) or [],
    }
    classification = normalize_vault_classification(
        classification,
        candidate_categories=list(hybrid_fallback["candidate_categories"]),
        fallback_primary=str(hybrid_fallback["fallback_primary"] or ""),
        fallback_confidence=hybrid_fallback["fallback_confidence"],
        fallback_secondary=list(hybrid_fallback["fallback_secondary"]),
    )
    classification = apply_text_message_override(
        classification,
        source_path=canonical_source_path or source_path,
        candidate_categories=list(hybrid_fallback["candidate_categories"]),
    )

    primary = str(classification.get("primary_label", "unknown") or "unknown")
    secondary = classification.get("secondary_labels", []) or []
    confidence_raw = classification.get("confidence", 0)
    summary = classification.get("summary", "")
    reason = classification.get("reason", "")
    sensitive_flags = classification.get("sensitive_flags", []) or []
    recommended_action = classification.get("recommended_action", "unknown")
    file_date_guess = classification.get("file_date_guess", "unknown")
    language = classification.get("language", "unknown")
    entity_summary = str(classification.get("entity_summary", "") or "")
    topic_summary = str(classification.get("topic_summary", "") or "")
    retrieval_topics = list(classification.get("retrieval_topics", []) or [])
    retrieval_terms = list(classification.get("retrieval_terms", []) or [])

    try:
        confidence = float(confidence_raw)
    except Exception:
        confidence = 0.0

    needs_review = confidence < 0.70 or primary == "unknown"
    summary = build_summary_fallback(
        summary=str(summary or ""),
        source_path=source_path,
        confidence=confidence,
        needs_review=needs_review,
    )
    reason = build_reason_fallback(
        reason=str(reason or ""),
        needs_review=needs_review,
        primary_label=primary,
    )
    category_parts = vault_category_parts(primary, secondary)
    category_label = vault_category_label(category_parts)
    category_path = Path(*category_parts)
    canonical_path_value = canonical_source_path or str(source_path)
    canonical_hash_value = canonical_source_hash or file_hash
    last_seen_filename_value = last_seen_filename or source_path.name

    if needs_review:
        note_dir = vault / "02 Needs Review"
    else:
        note_dir = vault / "01 Classified" / category_path

    note_dir.mkdir(parents=True, exist_ok=True)
    visible_source_name = display_source_name(
        source_path=source_path,
        canonical_source_path=canonical_path_value,
        last_seen_filename=last_seen_filename_value,
    )
    visible_title = Path(visible_source_name).stem
    existing_note_matches = _existing_generated_note_matches(
        vault,
        canonical_source_path=canonical_path_value,
        canonical_source_hash=canonical_hash_value,
        last_seen_filename=last_seen_filename_value,
    )
    preferred_note_path: Path | None = None
    preferred_note_metadata: dict[str, str] = {}
    duplicate_note_matches: list[tuple[int, Path, dict[str, str]]] = []
    if existing_note_matches:
        _, preferred_note_path, preferred_note_metadata = existing_note_matches[0]
        duplicate_note_matches = existing_note_matches[1:]

    note_existing_names = {path.name for path in note_dir.glob("*.md")}
    if preferred_note_path is not None and preferred_note_path.parent == note_dir:
        note_existing_names.discard(preferred_note_path.name)
    note_filename = build_note_filename(
        title=visible_title,
        primary_label=category_label,
        existing_names=note_existing_names,
    )
    note_path = note_dir / note_filename
    if preferred_note_path is not None and preferred_note_path != note_path:
        note_path.parent.mkdir(parents=True, exist_ok=True)
        preferred_note_path.replace(note_path)
        preferred_note_path = note_path
    elif preferred_note_path is not None:
        note_path = preferred_note_path

    extracted_link = ""
    if markdown is not None:
        extracted_dir = vault / "_system" / "extracted-markdown" / category_path
        extracted_dir.mkdir(parents=True, exist_ok=True)
        preferred_extracted_path = _resolve_note_attachment_path(
            vault,
            preferred_note_metadata,
            "extracted_markdown",
        )
        extracted_existing_names = {
            path.name for path in extracted_dir.glob("*.md")
        }
        if preferred_extracted_path is not None and preferred_extracted_path.parent == extracted_dir:
            extracted_existing_names.discard(preferred_extracted_path.name)
        extracted_path = extracted_dir / build_extracted_markdown_filename(
            title=visible_title,
            existing_names=extracted_existing_names,
        )
        if preferred_extracted_path is not None and preferred_extracted_path.exists():
            if preferred_extracted_path != extracted_path:
                extracted_path.parent.mkdir(parents=True, exist_ok=True)
                preferred_extracted_path.replace(extracted_path)
            else:
                extracted_path = preferred_extracted_path
        extracted_path.write_text(markdown, encoding="utf-8")
        extracted_link = f"[[{extracted_path.relative_to(vault).as_posix()}]]"

    source_link = build_canonical_source_link(canonical_source_path, visible_source_name)
    attachment_link = source_link
    attachment_mode = "canonical-source-link" if source_link else "none"

    # Prefer vault-local links when we are already attaching the original file.
    if attach_originals:
        attachment_link, attachment_mode = _build_vault_attachment_link(
            vault=vault,
            category_path=category_path,
            visible_source_name=visible_source_name,
            source_path=source_path,
        )
        source_link = attachment_link
    note_contract = build_note_contract_metadata(
        source_path=source_path,
        file_hash=file_hash,
        attachment_link=attachment_link,
        attachment_mode=attachment_mode,
        source_link=source_link,
        canonical_source_path=canonical_path_value,
        canonical_source_hash=canonical_hash_value,
        last_seen_filename=last_seen_filename_value,
    )

    tags = [f"classified/{obsidian_tag(primary)}"]

    for item in secondary:
        tags.append(f"classified/{obsidian_tag(item)}")

    if needs_review:
        tags.append("needs-review")

    for flag in sensitive_flags:
        if str(flag).lower() != "none":
            tags.append(f"sensitive/{obsidian_tag(flag)}")

    tags_yaml = "\n".join(f"  - {tag}" for tag in sorted(set(tags)))

    note_body = f"""---
type: classified-document
primary_label: {json.dumps(primary, ensure_ascii=False)}
secondary_labels: {yaml_list(secondary)}
confidence: {confidence}
source_file: {json.dumps(note_contract["canonical_source_path"], ensure_ascii=False)}
sha256: {json.dumps(file_hash)}
canonical_source_path: {json.dumps(note_contract["canonical_source_path"], ensure_ascii=False)}
canonical_source_hash: {json.dumps(note_contract["canonical_source_hash"], ensure_ascii=False)}
last_seen_filename: {json.dumps(note_contract["last_seen_filename"], ensure_ascii=False)}
attachment_mode: {json.dumps(note_contract["attachment_mode"], ensure_ascii=False)}
compatibility_attachment_path: {json.dumps(note_contract["compatibility_attachment_path"], ensure_ascii=False)}
source_link: {json.dumps(note_contract["source_link"], ensure_ascii=False)}
source_parser: {json.dumps(str(source_parser or ""), ensure_ascii=False)}
heuristic_primary_hint: {json.dumps(str(heuristic_primary_hint or ""), ensure_ascii=False)}
hybrid_live_source: {json.dumps(str(hybrid_live_source or ""), ensure_ascii=False)}
classified_at: {json.dumps(now_ak())}
sensitive_flags: {yaml_list(sensitive_flags)}
recommended_action: {json.dumps(recommended_action, ensure_ascii=False)}
attachment: {json.dumps(attachment_link, ensure_ascii=False)}
extracted_markdown: {json.dumps(extracted_link, ensure_ascii=False)}
tags:
{tags_yaml}
---

# {visible_source_name}

## Summary

{summary}

## Original File

{attachment_link if attachment_link else "Original file was not copied into the vault. Re-run with `--attach-originals` if desired."}

## Extracted Markdown File

{extracted_link if extracted_link else "No extracted Markdown file was written."}

## Extracted Markdown Preview

{markdown[:20000] if markdown else "_No Markdown extraction available for this file._"}

## Classification

| Field | Value |
|---|---|
| Primary label | `{primary}` |
| Secondary labels | `{", ".join(map(str, secondary))}` |
| Confidence | `{confidence}` |
| Sensitive flags | `{", ".join(map(str, sensitive_flags))}` |
| Recommended action | `{recommended_action}` |
| File date guess | `{file_date_guess}` |
| Language | `{language}` |
| SHA-256 | `{file_hash}` |

## Retrieval

| Field | Value |
|---|---|
| Topics | `{", ".join(map(str, retrieval_topics)) or topic_summary or "none"}` |
| Entities | `{entity_summary or "none"}` |
| Retrieval terms | `{", ".join(map(str, retrieval_terms)) or "none"}` |

## Reason

{reason}

## System Metadata

| Field | Value |
|---|---|
| Attachment mode | `{note_contract["attachment_mode"]}` |
| Compatibility attachment path | `{note_contract["compatibility_attachment_path"] or "none"}` |
| Source parser | `{str(source_parser or "") or "unknown"}` |
| Heuristic primary hint | `{str(heuristic_primary_hint or "") or "unknown"}` |
| Hybrid live source | `{str(hybrid_live_source or "") or "unknown"}` |
"""

    note_path.write_text(note_body, encoding="utf-8")
    for _, duplicate_note_path, duplicate_note_metadata in duplicate_note_matches:
        if duplicate_note_path == note_path:
            continue
        _remove_duplicate_generated_note(vault, duplicate_note_path, duplicate_note_metadata)
    return note_path

def write_index(
    vault: Path,
    notes: List[Path],
    note_records: Optional[List[Dict[str, Any]]] = None,
) -> None:
    index = vault / "Classification Index.md"
    topic_counts: Counter[str] = Counter()
    entity_counts: Counter[str] = Counter()

    for record in note_records or []:
        for topic in record.get("retrieval_topics", []) or []:
            normalized = str(topic or "").strip()
            if normalized:
                topic_counts[normalized] += 1
        entity_summary = str(record.get("entity_summary", "") or "")
        for fragment in [part.strip() for part in entity_summary.split(";") if part.strip()]:
            entity_counts[fragment] += 1

    lines = [
        "---",
        "type: classification-index",
        "system: local-document-classifier",
        "---",
        "",
        "# Classification Index",
        "",
        f"Last updated: {now_ak()}",
        "",
        "## Discovery topics",
        "",
    ]

    if topic_counts:
        for topic, count in topic_counts.most_common(20):
            lines.append(f"- `{topic}` ({count})")
    else:
        lines.append("- No retrieval topics recorded yet.")

    lines.extend(
        [
            "",
            "## Discovery entities",
            "",
        ]
    )

    if entity_counts:
        for entity, count in entity_counts.most_common(20):
            lines.append(f"- {entity} ({count})")
    else:
        lines.append("- No retrieval entities recorded yet.")

    lines.extend(
        [
            "",
            "## Recent notes",
            "",
        ]
    )

    for note in notes[-100:]:
        rel = note.relative_to(vault).as_posix()
        lines.append(f"- [[{rel}]]")

    index.write_text("\n".join(lines) + "\n", encoding="utf-8")

def main() -> int:
    parser = argparse.ArgumentParser(description="Dockerized local document classifier that writes to an Obsidian vault.")
    parser.add_argument("input", nargs="?", default="/input", help="Input file or folder inside the container. Default: /input")
    parser.add_argument("--vault", default="/vault", help="Obsidian vault path inside the container. Default: /vault")
    parser.add_argument("--output", default="/output", help="Output folder for manifest/errors. Default: /output")
    parser.add_argument("--work-dir", default="/tmp/work", help="Temporary work directory.")
    parser.add_argument("--ollama-url", default=os.environ.get("OLLAMA_URL", "http://ollama:11434"))
    parser.add_argument("--model", default=os.environ.get("CLASSIFY_MODEL", "qwen2.5:3b"))
    parser.add_argument("--vision-model", default=os.environ.get("VISION_MODEL", "qwen2.5vl:3b"))
    parser.add_argument("--categories", default="__AUTO__")
    parser.add_argument("--max-chars", type=int, default=16000)
    parser.add_argument("--attach-originals", action="store_true")
    parser.add_argument("--no-vision", action="store_true")
    parser.add_argument("--enable-codex-arbiter", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--timing-output", default="", help="Optional JSON file path for per-run timing output.")
    parser.add_argument("--process-shadow-queue", action="store_true")
    parser.add_argument("--retrain-hybrid-model", action="store_true")
    parser.add_argument("--write-readiness-report", action="store_true")
    parser.add_argument("--canonical-source-path", default="")
    parser.add_argument("--canonical-source-hash", default="")
    parser.add_argument("--last-seen-filename", default="")

    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    vault = Path(args.vault).resolve()
    output = Path(args.output).resolve()
    work_dir = Path(args.work_dir).resolve()

    if args.categories == "__AUTO__":
        categories = load_categories()
    else:
        categories = [x.strip() for x in args.categories.split(",") if x.strip()]

    ensure_vault(vault)
    output.mkdir(parents=True, exist_ok=True)
    work_dir.mkdir(parents=True, exist_ok=True)

    if args.process_shadow_queue:
        wait_for_ollama(
            args.ollama_url,
            required_models=[args.model, args.vision_model],
        )
        result = process_shadow_queue_command(
            categories=categories,
            ollama_url=args.ollama_url,
            model=args.model,
            vision_model=args.vision_model,
            max_chars=args.max_chars,
        )
        print(json.dumps(result, indent=2))
        return 0 if result.get("ok") else 1

    if args.write_readiness_report:
        result = write_readiness_report(gating_config=load_hybrid_gating_config())
        print(json.dumps(result, indent=2))
        return 0 if result.get("ok") else 1

    if args.retrain_hybrid_model:
        result = ensure_lightgbm_model()
        if not result.get("ok") or not result.get("created"):
            retrain = maybe_retrain_from_shadow_data(min_rows=3)
            print(json.dumps(retrain, indent=2))
            return 0 if retrain.get("retrained") else 1
        print(json.dumps(result, indent=2))
        return 0

    if args.self_test:
        wait_for_ollama(
            args.ollama_url,
            required_models=[args.model, args.vision_model],
        )
        print(json.dumps({
            "ok": True,
            "time_alaska": now_ak(),
            "ollama_url": args.ollama_url,
            "vault": str(vault),
            "output": str(output),
            "model": args.model,
            "vision_model": args.vision_model,
            "category_count": len(categories),
        }, indent=2))
        return 0

    if not input_path.exists():
        print(f"[ERROR] Input path does not exist inside container: {input_path}", file=sys.stderr)
        return 2

    files = list(iter_input_files(input_path))

    required_models = [args.model]
    if args.process_shadow_queue or (
        not args.no_vision and any(path.suffix.lower() in IMAGE_EXTENSIONS for path in files)
    ):
        required_models.append(args.vision_model)

    wait_for_ollama(
        args.ollama_url,
        required_models=required_models,
    )

    if not files:
        print(f"[WARN] No supported files found under: {input_path}")
        return 0

    ensure_lightgbm_model()

    manifest = output / "manifest.jsonl"
    notes: List[Path] = []
    note_records: List[Dict[str, Any]] = []
    successes = 0
    failures = 0
    timing_records: List[Dict[str, Any]] = []
    gating_config = load_hybrid_gating_config()
    heuristic_rules = load_heuristic_rules()

    with manifest.open("a", encoding="utf-8") as mf:
        for source_path in files:
            print(f"[INFO] Classifying: {source_path}")
            markdown: Optional[str] = None
            file_hash = ""
            hybrid_meta: Optional[Dict[str, Any]] = None
            heuristic_classification: Optional[Dict[str, Any]] = None
            file_started_at = time.perf_counter()
            ext = source_path.suffix.lower()
            timing: Dict[str, Any] = {
                "source_path": str(source_path),
                "filename": source_path.name,
                "extension": ext,
                "mode": "image" if ext in IMAGE_EXTENSIONS and not args.no_vision else "document",
                "attach_originals": args.attach_originals,
                "vision_enabled": not args.no_vision,
                "codex_arbiter_enabled": bool(args.enable_codex_arbiter),
            }

            try:
                hash_started_at = time.perf_counter()
                file_hash = sha256_file(source_path)
                timing["sha256_ms"] = elapsed_ms(hash_started_at)

                if ext in IMAGE_EXTENSIONS and not args.no_vision:
                    classification, hybrid_meta, markdown = classify_image(
                        source_path=source_path,
                        reviewed_source_path=args.canonical_source_path or None,
                        categories=categories,
                        ollama_url=args.ollama_url,
                        model=args.model,
                        vision_model=args.vision_model,
                        max_chars=args.max_chars,
                        gating_config=gating_config,
                        heuristic_rules=heuristic_rules,
                        timing=timing,
                    )
                    timing["parser"] = (
                        f"image-ocr-{classification.get('ocr_engine', '')}".rstrip("-")
                        if hybrid_meta is not None
                        else "vision-qwen"
                    )
                    timing["markdown_chars"] = len(markdown or "")
                    timing["clipped_markdown_chars"] = len((markdown or "")[: args.max_chars])
                    timing["classifier"] = "ocr-document-path" if hybrid_meta is not None else "vision-qwen-fallback"
                elif ext in SPREADSHEET_EXTENSIONS:
                    parse_started_at = time.perf_counter()
                    markdown, parser_name, spreadsheet_metadata = parse_spreadsheet_fast(source_path)
                    timing["parse_ms"] = elapsed_ms(parse_started_at)
                    timing["parser"] = parser_name
                    timing["candidate_category_count"] = len(
                        [
                            label
                            for label in ["spreadsheet", "financial", "work", "report", "tax", "legal", "medical", "insurance", "technical", "unknown", "needs-review"]
                            if label in categories
                        ]
                    )
                    timing["markdown_chars"] = len(markdown)
                    timing["clipped_markdown_chars"] = len(markdown)
                    timing["sheet_count"] = spreadsheet_metadata.get("sheet_count", 0)
                    extraction_metadata = {
                        "ocr_engine": "",
                        "ocr_quality": "",
                        "ocr_char_count": 0,
                        "extraction_quality": extraction_quality_for_text(markdown),
                    }
                    timing["extraction_quality"] = extraction_metadata["extraction_quality"]
                    markdown, heuristic_classification, spreadsheet_metadata = classify_spreadsheet_fast(
                        source_path=source_path,
                        categories=categories,
                        markdown=markdown,
                        metadata=spreadsheet_metadata,
                    )
                    classification, hybrid_meta = resolve_hybrid_document_decision(
                        source_path=source_path,
                        reviewed_source_path=args.canonical_source_path or None,
                        markdown=markdown,
                        parser_name=parser_name,
                        categories=categories,
                        heuristic_result=heuristic_classification,
                        ollama_url=args.ollama_url,
                        model=args.model,
                        max_chars=args.max_chars,
                        gating_config=gating_config,
                        timing=timing,
                        extraction_metadata=extraction_metadata,
                    )
                    classification["extraction_quality"] = extraction_metadata["extraction_quality"]
                    timing["classifier"] = (
                        "heuristic-spreadsheet-fast-path"
                        if hybrid_meta["decision"]["live_source"] == "heuristic-fast-path"
                        else "taxonomy-aware-inline-llm"
                    )
                    if hybrid_meta["decision"]["live_source"] == "heuristic-fast-path":
                        timing["model_ms"] = 0.0
                else:
                    parse_started_at = time.perf_counter()
                    markdown, parser_name, extraction_metadata = parse_document(source_path, work_dir)
                    timing["parse_ms"] = elapsed_ms(parse_started_at)
                    timing["parser"] = parser_name
                    timing["extraction_quality"] = extraction_metadata.get("extraction_quality", "")
                    if extraction_metadata.get("ocr_engine"):
                        timing["ocr_engine"] = extraction_metadata.get("ocr_engine", "")
                        timing["ocr_quality"] = extraction_metadata.get("ocr_quality", "")
                        timing["ocr_chars"] = int(extraction_metadata.get("ocr_char_count", 0) or 0)
                    heuristic_classification = classify_document_fast(
                        source_path=source_path,
                        markdown=markdown,
                        categories=categories,
                        rules=heuristic_rules,
                    )
                    timing["markdown_chars"] = len(markdown)
                    timing["clipped_markdown_chars"] = len(markdown[:args.max_chars])
                    classification, hybrid_meta = resolve_hybrid_document_decision(
                        source_path=source_path,
                        reviewed_source_path=args.canonical_source_path or None,
                        markdown=markdown,
                        parser_name=parser_name,
                        categories=categories,
                        heuristic_result=heuristic_classification,
                        ollama_url=args.ollama_url,
                        model=args.model,
                        max_chars=args.max_chars,
                        gating_config=gating_config,
                        timing=timing,
                        extraction_metadata=extraction_metadata,
                    )
                    classification["extraction_quality"] = extraction_metadata.get("extraction_quality", "")
                    if extraction_metadata.get("ocr_engine"):
                        classification["ocr_engine"] = extraction_metadata.get("ocr_engine", "")
                        classification["ocr_quality"] = extraction_metadata.get("ocr_quality", "")
                        classification["ocr_char_count"] = int(extraction_metadata.get("ocr_char_count", 0) or 0)
                    timing["classifier"] = (
                        "heuristic-document-fast-path"
                        if hybrid_meta["decision"]["live_source"] == "heuristic-fast-path" and heuristic_classification is not None
                        else "taxonomy-aware-inline-llm"
                    )
                    if hybrid_meta["decision"]["live_source"] == "heuristic-fast-path":
                        timing["model_ms"] = 0.0

                fallback_primary = str(
                    ((hybrid_meta or {}).get("decision") or {}).get("selected_primary_hint")
                    or ((hybrid_meta or {}).get("lightgbm") or {}).get("top_label")
                    or (heuristic_classification or {}).get("primary_label")
                    or ""
                )
                fallback_confidence = float(
                    (((hybrid_meta or {}).get("lightgbm") or {}).get("top_probability") or 0.0)
                )
                classification = normalize_vault_classification(
                    classification,
                    candidate_categories=list(
                        ((hybrid_meta or {}).get("taxonomy_candidates") or classification.get("candidate_categories_used", []) or [])
                    ),
                    fallback_primary=fallback_primary,
                    fallback_confidence=fallback_confidence,
                    fallback_secondary=[str((heuristic_classification or {}).get("primary_label") or "")],
                )
                hybrid_live_source = str(((hybrid_meta or {}).get("decision") or {}).get("live_source", "") or "")
                classification, hybrid_live_source, codex_arbiter_meta = apply_codex_arbiter_if_enabled(
                    enabled=bool(args.enable_codex_arbiter),
                    source_path=source_path,
                    markdown=markdown or "",
                    local_classification=classification,
                    heuristic_classification=heuristic_classification,
                    hybrid_live_source=hybrid_live_source,
                )
                timing["codex_arbiter_status"] = codex_arbiter_meta.get("status", "disabled")
                timing["codex_arbiter_applied"] = bool(codex_arbiter_meta.get("applied"))
                if codex_arbiter_meta.get("duration_ms") is not None:
                    timing["codex_arbiter_ms"] = codex_arbiter_meta.get("duration_ms")
                classification = apply_text_message_override(
                    classification,
                    source_path=args.canonical_source_path or source_path,
                    candidate_categories=list(
                        ((hybrid_meta or {}).get("taxonomy_candidates") or classification.get("candidate_categories_used", []) or [])
                    ),
                )
                timing["final_live_source"] = hybrid_live_source
                retrieval_text_source = "\n".join(
                    filter(
                        None,
                        [
                            markdown or "",
                            str(classification.get("summary", "") or ""),
                            str(classification.get("reason", "") or ""),
                        ],
                    )
                )
                retrieval_metadata = build_retrieval_metadata(
                    source_path=source_path,
                    text=retrieval_text_source,
                    classification=classification,
                )
                classification.update(retrieval_metadata)
                note_started_at = time.perf_counter()
                note_path = write_obsidian_note(
                    vault=vault,
                    source_path=source_path,
                    file_hash=file_hash,
                    markdown=markdown,
                    classification=classification,
                    attach_originals=args.attach_originals,
                    canonical_source_path=args.canonical_source_path or None,
                    canonical_source_hash=args.canonical_source_hash or None,
                    last_seen_filename=args.last_seen_filename or None,
                    source_parser=str(timing.get("parser", "") or ""),
                    heuristic_primary_hint=str((heuristic_classification or {}).get("primary_label", "") or "unknown"),
                    hybrid_live_source=hybrid_live_source,
                )
                timing["note_write_ms"] = elapsed_ms(note_started_at)
                timing["primary_label"] = classification.get("primary_label", "unknown")
                timing["secondary_label_count"] = len(classification.get("secondary_labels", []) or [])
                timing["confidence"] = classification.get("confidence")
                timing["ok"] = True
                timing["total_ms"] = elapsed_ms(file_started_at)
                visible_source_name = display_source_name(
                    source_path=source_path,
                    canonical_source_path=args.canonical_source_path or None,
                    last_seen_filename=args.last_seen_filename or None,
                )
                record_category_path = "/".join(
                    vault_category_parts(
                        str(classification.get("primary_label", "unknown") or "unknown"),
                        classification.get("secondary_labels", []) or [],
                    )
                )
                record_source_link = build_canonical_source_link(
                    args.canonical_source_path or None,
                    visible_source_name,
                )
                record_attachment_link = record_source_link
                record_attachment_mode = "canonical-source-link" if record_source_link else "none"
                if args.attach_originals:
                    record_attachment_link = (
                        f"[[90 Attachments/"
                        f"{record_category_path}/"
                        f"{visible_source_name}]]"
                    )
                    record_attachment_mode = "copied-compatibility"
                    record_source_link = record_attachment_link

                record = {
                    "ok": True,
                    "classified_at": now_ak(),
                    "source_path": str(source_path),
                    "sha256": file_hash,
                    "note_path": str(note_path),
                    **build_note_contract_metadata(
                        source_path=source_path,
                        file_hash=file_hash,
                        attachment_link=record_attachment_link,
                        attachment_mode=record_attachment_mode,
                        source_link=record_source_link,
                        canonical_source_path=args.canonical_source_path or None,
                        canonical_source_hash=args.canonical_source_hash or None,
                        last_seen_filename=args.last_seen_filename or None,
                    ),
                    "entity_summary": retrieval_metadata["entity_summary"],
                    "topic_summary": retrieval_metadata["topic_summary"],
                    "retrieval_terms": retrieval_metadata["retrieval_terms"],
                    "retrieval_text": retrieval_metadata["retrieval_text"],
                    "classification": classification,
                    "hybrid": hybrid_meta,
                    "timing": timing,
                }

                manifest_started_at = time.perf_counter()
                mf.write(json.dumps(record, ensure_ascii=False) + "\n")
                mf.flush()
                timing["manifest_write_ms"] = elapsed_ms(manifest_started_at)
                timing["total_ms"] = elapsed_ms(file_started_at)

                if should_enqueue_shadow_job(gating_config, (hybrid_meta or {}).get("decision", {}).get("live_source", "")):
                    shadow_payload = {
                        "mode": "image" if ext in IMAGE_EXTENSIONS and not args.no_vision else "document",
                        "filename": source_path.name,
                        "extension": ext,
                        "parser": timing.get("parser"),
                        "source_path": str(source_path),
                        "markdown": markdown or "",
                        "heuristic_result": heuristic_classification,
                        "lightgbm_result": (hybrid_meta or {}).get("lightgbm"),
                        "live_result": classification,
                        "live_source": (hybrid_meta or {}).get("decision", {}).get("live_source", ""),
                        "taxonomy_candidates": (hybrid_meta or {}).get("taxonomy_candidates", []),
                        "entity_summary": retrieval_metadata["entity_summary"],
                        "topic_summary": retrieval_metadata["topic_summary"],
                        "retrieval_terms": retrieval_metadata["retrieval_terms"],
                        "retrieval_text": retrieval_metadata["retrieval_text"],
                        "ocr_engine": classification.get("ocr_engine", ""),
                        "ocr_quality": classification.get("ocr_quality", ""),
                        "ocr_char_count": int(classification.get("ocr_char_count", 0) or 0),
                        "extraction_quality": classification.get("extraction_quality", ""),
                        "text_preview": ((markdown or classification.get("summary", ""))[:12000]),
                    }
                    enqueue_shadow_job(shadow_payload)
                    timing["shadow_enqueued"] = True

                notes.append(note_path)
                note_records.append(
                    {
                        "note_path": str(note_path),
                        **retrieval_metadata,
                    }
                )
                timing_records.append(dict(timing))
                successes += 1

                print(f"[OK] {source_path.name} => {classification.get('primary_label', 'unknown')} note={note_path}")

            except Exception as e:
                failures += 1
                timing["ok"] = False
                timing["error"] = str(e)
                timing["total_ms"] = elapsed_ms(file_started_at)

                record = {
                    "ok": False,
                    "classified_at": now_ak(),
                    "source_path": str(source_path),
                    "sha256": file_hash,
                    "error": str(e),
                    "timing": timing,
                }

                mf.write(json.dumps(record, ensure_ascii=False) + "\n")
                mf.flush()
                timing_records.append(dict(timing))

                print(f"[FAIL] {source_path}: {e}", file=sys.stderr)

    write_index(vault, notes, note_records)
    write_readiness_report(gating_config=gating_config)

    if args.timing_output:
        timing_output = Path(args.timing_output)
        timing_output.parent.mkdir(parents=True, exist_ok=True)
        payload: Dict[str, Any]
        if len(timing_records) == 1:
            payload = dict(timing_records[0])
        else:
            payload = {
                "ok": failures == 0,
                "successes": successes,
                "failures": failures,
                "file_count": len(timing_records),
                "files": timing_records,
            }
        timing_output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"[DONE] successes={successes} failures={failures}")
    print(f"[DONE] vault={vault}")
    print(f"[DONE] manifest={manifest}")

    return 0 if failures == 0 else 1

if __name__ == "__main__":
    raise SystemExit(main())
