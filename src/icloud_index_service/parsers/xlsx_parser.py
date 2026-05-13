from __future__ import annotations

from io import BytesIO


def extract_text_from_xlsx_bytes(payload: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(payload), data_only=True)
    parts: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                parts.append(str(cell))
    return "\n".join(parts).strip()
